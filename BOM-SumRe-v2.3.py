
# v2.0版本说明
# 1、使用多维数组操作
# 2、删除无数据的列
# 3、增加数据源列
# v2.2版本说明
# 1、解决数据源列、时间列不整齐和日期格式显示的问题
# 2、取消无效数据列，增加取不连续区域内最大值的列
# v2.3版本说明
# 1、解决M价格列数据不整齐的问题；
# 2、根据料号分类
# 3、增加未去重表
import os
import xlrd
from openpyxl import Workbook
from tqdm import tqdm
from datetime import datetime
from datetime import date
from openpyxl.styles import Font
from openpyxl.utils import get_column_letter
from openpyxl.utils.dataframe import dataframe_to_rows
import pandas as pd
import frozen_dir

def savelog(file,log):
    dirname = os.path.dirname(file)
    if not os.path.exists(dirname):
        os.mkdir(dirname)
    with open(file,'a',encoding='utf-8') as f:
        f.write(log+'\n')

def create_or_load_workbook(filename):
    # 如果文件存在，则先删除
    if os.path.exists(filename):
        os.remove(filename)
    
    # 创建新的工作簿并返回
    wb = Workbook()
    wb.active.title = "Sum"
    wb.save(filename)
    return wb

def read_existing_data(ws_summary):
    # 从工作表中读取已存在的数据并返回多维数组
    data = []
    for row in ws_summary.iter_rows(values_only=True):
        data.append([float(cell) if isinstance(cell, (int, float)) else cell for cell in row])
    return data

def remove_empty_columns(data):
    if not data:
        return []
    
    # 找到所有行中相同索引号都为空的列的索引
    empty_column_idxs = set()
    num_columns = len(data[0])  # 确定数据中列的数量
    for idx in range(num_columns):  # 遍历所有列
        column_values = [row[idx] for row in data if len(row) > idx]  # 提取当前列的所有值
        if all(value is None or value == '' for value in column_values):  # 如果当前列的所有值都为空或None
            empty_column_idxs.add(idx)  # 将当前列的索引添加到空列索引集合中

    # 创建一个新的数据列表，用于保存删除空列后的数据
    new_data = []
    # 删除所有行中相同索引号都为空的列
    for row in data:
        # 仅保留不在空列索引集合中的列
        new_row = [cell for idx, cell in enumerate(row) if idx not in empty_column_idxs]
        new_data.append(new_row)
    return new_data

def set_column_width_and_font(ws):
    # 设置字体
    for row in ws.iter_rows():
        for cell in row:
            cell.font = Font(name='細明體', size=9)

    # 自适应列宽
    for column_cells in ws.columns:
        max_length = 0
        for cell in column_cells:
            if cell.value is not None:
                max_length = max(max_length, len(str(cell.value)))
        if max_length < 14:
            adjusted_width = max_length
        else:
            adjusted_width = max_length * 0.7
        column_letter = get_column_letter(column_cells[0].column)
        ws.column_dimensions[column_letter].width = adjusted_width

def find_in_matrix(matrix, target):
    for i in range(len(matrix)):
        for j in range(len(matrix[0])):
            if matrix[i][j] == target :
                return(i,j)
    return None

def update_data_with_max_column(data):
    # 步骤 1：找到包含“料件编号”的行索引
    material_code_row_idx = None
    for idx, row in enumerate(data):
        if "料件编号" in row:
            material_code_row_idx = idx
            break
    
    if material_code_row_idx is None:
        return data

    # 步骤 2：找到“BOM单位”的索引，并在其后插入“M价格”
    bom_unit_idx = data[material_code_row_idx].index("BOM单位")
    for idx, row in enumerate(data):
        if idx != material_code_row_idx:  # 跳过标题行
            row.insert(bom_unit_idx+1, None)  # 在 M价格 列之前插入空列

    # 步骤 3：更新所有行的数据，并获取“采购单价”、“平均单价”、“市价”的索引
    purchase_price_idx = None
    average_price_idx = None
    market_price_idx = None

    for row in data:
        if "M价格" in row:
            m_idx = row.index("M价格")
        if "采购单价" in row:
            purchase_price_idx = row.index("采购单价")
        if "平均单价" in row:
            average_price_idx = row.index("平均单价")
        if "市价" in row:
            market_price_idx = row.index("市价")

    for idx, row in enumerate(data[material_code_row_idx:]):  # 修正遍历行时的索引起点
        if idx != material_code_row_idx:
            # 获取采购单价、平均单价、市价，确保正确识别数字类型的值
            purchase_price = float(row[purchase_price_idx]) if isinstance(row[purchase_price_idx], (int, float)) else 0
            average_price = float(row[average_price_idx]) if isinstance(row[average_price_idx], (int, float)) else 0
            market_price = float(row[market_price_idx]) if isinstance(row[market_price_idx], (int, float)) else 0
            # 计算 M价格
            m_price = max(purchase_price, average_price, market_price)
            row[bom_unit_idx+1] = m_price  # 更新 M价格 列的值
        
    return data
def categorize_data_by_material_code(data, material_code_column_idx, ws_summary):
    remove_data = [] #存储去重后的物料数据
    cta_board_data = []  # 存储20CT开头的数据
    nsb_board_data = []  # 存储20SB开头的数据
    nsk_board_data = []  # 存储20Sk开头的数据
    ddr_data = []  # 存储72开头的数据
    ssd_data = [] # 存储73开头的数据
    power_data = []  # 存储74开头的数据
    other_data = [] #存储剩余数据

    for row in data:
        material_code = row[material_code_column_idx]
        if len(material_code) == 13 or material_code=='料件编号' : 
            if material_code in [existing_row[material_code_column_idx] for existing_row in remove_data]:
                # 如果料件编号已存在，则覆盖原有内容并更新时间和文件名信息
                for existing_row in remove_data:
                    if existing_row[material_code_column_idx] == material_code:
                        # existing_row[:len(row_values)] = row_values
                        # existing_row[-2:] = [modification_time, os.path.basename(file_path)]
                        break
            else:
                # 如果料件编号不存在于已存在数据中，则添加到工作表中
                remove_data.append(row)
    for row in remove_data:
        material_code = row[material_code_column_idx]
        if ((datetime.strptime(str(date.today()), "%Y-%m-%d")-datetime.strptime(str(row[-2]), "%Y-%m-%d")).days) > 40 :
            
            print(row[-1],(datetime.strptime(str(date.today()), "%Y-%m-%d")-datetime.strptime(str(row[-2]), "%Y-%m-%d")).days)
        if isinstance(material_code, str) and len(material_code) == 13 or material_code=='料件编号':
            if material_code == '料件编号':
                cta_board_data.append(row)
                nsb_board_data.append(row)
                nsk_board_data.append(row)
                ddr_data.append(row)
                ssd_data.append(row)
                power_data.append(row)
                other_data.append(row)
            elif material_code.startswith("20CT")  :
                cta_board_data.append(row)
            elif material_code.startswith("20SB"):
                nsb_board_data.append(row)
            elif material_code.startswith("20SK") :
                nsk_board_data.append(row)
            elif material_code.startswith("72"):
               ddr_data.append(row)
            elif material_code.startswith("73"):
               ssd_data.append(row)
            elif material_code.startswith("74"):
               power_data.append(row)
            else:
                other_data.append(row)

    # 创建或获取CTA主板和NSB主板表
    remove_ws = ws_summary.parent.create_sheet(title="reData")
    cta_board_ws = ws_summary.parent.create_sheet(title="CTA主板")
    nsb_board_ws = ws_summary.parent.create_sheet(title="NSB主板")
    nsk_board_ws = ws_summary.parent.create_sheet(title="NSK转卡")
    ddr_ws = ws_summary.parent.create_sheet(title="内存")
    ssd_ws = ws_summary.parent.create_sheet(title="SSD")
    power_ws = ws_summary.parent.create_sheet(title="电源")
    other_ws = ws_summary.parent.create_sheet(title="其他")
    # 将数据写入不同的表
    for sheet, data in zip([remove_ws,cta_board_ws, nsb_board_ws,nsk_board_ws,ddr_ws,ssd_ws,power_ws,other_ws], [remove_data,cta_board_data, nsb_board_data,nsk_board_data,ddr_data,ssd_data,power_data,other_data]):
        for row in data:
            sheet.append(row)

    # 设置列宽和字体
    set_column_width_and_font(remove_ws)
    set_column_width_and_font(cta_board_ws)
    set_column_width_and_font(nsb_board_ws)
    set_column_width_and_font(nsk_board_ws)
    set_column_width_and_font(ddr_ws)
    set_column_width_and_font(ssd_ws)
    set_column_width_and_font(power_ws)
    set_column_width_and_font(other_ws)

def main():
    # 获取当前脚本路径
    current_script_path = os.path.dirname(os.path.abspath(__file__))
    
    # 获取BOMData文件夹路径
    bomdata_folder = os.path.join(current_script_path,'BOMData')
    
    # 获取BOMSummary&Remove.xlsx文件路径
    bomsummary_path = os.path.join(current_script_path,"..", 'BOM_Sum.xlsx')

    # 加载或创建BOMSummary&Remove.xlsx文件
    wb_summary = create_or_load_workbook(bomsummary_path)
    ws_summary = wb_summary.active

    # 读取已存在的数据并记录料件编号及行号
    existing_data = read_existing_data(ws_summary)

    # 获取BOMData文件夹下所有.xls文件的路径
    xls_files = [os.path.join(bomdata_folder, f) for f in os.listdir(bomdata_folder) if f.endswith('.xls')]

    # 进度条设置
    total_files = len(xls_files)
    with tqdm(total=total_files, desc="Processing files", ncols=100) as pbar:
        for idx, file_path in enumerate(xls_files):
            # 打开.xls文件
            workbook = xlrd.open_workbook(file_path)
            worksheet = workbook.sheet_by_index(0)

            # 获取文件的修改日期作为时间戳
            modification_time = datetime.fromtimestamp(os.path.getmtime(file_path)).strftime('%Y-%m-%d') #%H:%M:%S
            # 遍历.xls文件的每一行
            for row_idx in range(worksheet.nrows):
                row_values = worksheet.row_values(row_idx)
                material_code = row_values[1]  # 假设料件编号在第二列

                # 检查料件编号长度或值
                if len(str(material_code)) == 13 or material_code == '料件编号':
                        existing_data.append(row_values + [modification_time, os.path.basename(file_path)])
            # 更新进度条
            pbar.update(1)

    # 删除数组中所有行中相同索引号都为空的列
    existing_data = remove_empty_columns(existing_data)
    # 增加max
    updated_data  = update_data_with_max_column(existing_data)

    # 获取料件编号所在列的索引
    material_code_column_idx = None
    for idx, column_name in enumerate(updated_data[0]):
        if column_name == "料件编号":
            material_code_column_idx = idx
            break
    # 分类数据到不同的表
    categorize_data_by_material_code(updated_data, material_code_column_idx, ws_summary)

    # 清空工作表并写入新数据
    ws_summary.delete_rows(1, ws_summary.max_row)
    for row_data in updated_data:
        ws_summary.append(row_data)

    # 设置列宽和字体
    set_column_width_and_font(ws_summary)

    # 保存BOMSummary&Remove.xlsx文件
    wb_summary.save(bomsummary_path)

if __name__ == "__main__":
    main()
    file = frozen_dir.app_path()+r'\log\RunLog.log'
    # savelog(file,datetime.strptime(str(date.today()), "%Y-%m-%d"))
    # savelog(file,"test")
    print(file)
    savelog(file,'hello you')
    savelog(file,'文件路径%s' %(file))